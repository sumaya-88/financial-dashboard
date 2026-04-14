"""
Financial Performance Dashboard - KPI Analysis & Excel Model Builder
Runs full analysis and exports a multi-sheet Excel workbook
"""

import pandas as pd
import numpy as np
import json
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

DATA_DIR = "/home/claude/financial-dashboard/data"
OUT_PATH = "/home/claude/financial-dashboard/Financial_Dashboard.xlsx"

YEARS = [2020, 2021, 2022, 2023, 2024]
MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
QUARTERS = ["Q1","Q2","Q3","Q4"]

# ── Load data ──────────────────────────────────────────────────────────────────
df_a = pd.read_csv(f"{DATA_DIR}/annual_pnl.csv")
df_m = pd.read_csv(f"{DATA_DIR}/monthly_revenue.csv")
df_as = pd.read_csv(f"{DATA_DIR}/asset_quality.csv")
df_s = pd.read_csv(f"{DATA_DIR}/segment_revenue.csv")
df_q = pd.read_csv(f"{DATA_DIR}/quarterly_kpis.csv")

# ── Style helpers ──────────────────────────────────────────────────────────────
DARK_BLUE  = "0D1B2A"
MID_BLUE   = "1B3A5C"
ACCENT     = "1E88E5"
LIGHT_BLUE = "BBDEFB"
GREEN      = "1B8A4E"
RED        = "C62828"
GOLD       = "F9A825"
WHITE      = "FFFFFF"
LIGHT_GREY = "F5F7FA"
BORDER_CLR = "BDBDBD"

def hdr(ws, row, col, value, dark=True):
    cell = ws.cell(row=row, column=col, value=value)
    bg = DARK_BLUE if dark else MID_BLUE
    cell.font = Font(bold=True, color=WHITE, size=11, name="Arial")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=WHITE)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def data_cell(ws, row, col, value, fmt=None, color=None, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", size=10, bold=bold,
                     color=color if color else "000000")
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if fmt:
        cell.number_format = fmt
    if row % 2 == 0:
        cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
    thin = Side(style="thin", color=BORDER_CLR)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def label_cell(ws, row, col, value, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", size=10, bold=bold)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if row % 2 == 0:
        cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
    thin = Side(style="thin", color=BORDER_CLR)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def section_title(ws, row, col, value, span=8):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, size=13, name="Arial", color=DARK_BLUE)
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)

def set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

def set_row_height(ws, rows, height=18):
    for r in rows:
        ws.row_dimensions[r].height = height

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – Executive Summary
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_view.showGridLines = False

# Title banner
ws1.merge_cells("A1:L2")
title_cell = ws1["A1"]
title_cell.value = "APEX BANK — FINANCIAL PERFORMANCE DASHBOARD"
title_cell.font = Font(bold=True, size=18, name="Arial", color=WHITE)
title_cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 10

ws1.merge_cells("A3:L3")
sub_cell = ws1["A3"]
sub_cell.value = "Annual KPI Summary  |  FY 2020–2024  |  All figures in USD Millions"
sub_cell.font = Font(size=10, name="Arial", color=WHITE, italic=True)
sub_cell.fill = PatternFill("solid", fgColor=MID_BLUE)
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[3].height = 18

# KPI Header row
cols = ["Metric", "2020", "2021", "2022", "2023", "2024", "YoY Δ", "5Y CAGR"]
for ci, label in enumerate(cols, 1):
    hdr(ws1, 5, ci, label)

# Metrics to display
metrics = [
    ("Total Revenue ($M)",       "Total_Revenue_M",       '#,##0.0',  False),
    ("  Interest Income ($M)",   "Interest_Income_M",     '#,##0.0',  False),
    ("  Fee Income ($M)",        "Fee_Income_M",          '#,##0.0',  False),
    ("Operating Expenses ($M)",  "Operating_Expenses_M",  '#,##0.0',  False),
    ("Provision for Losses ($M)","Provision_Losses_M",    '#,##0.0',  False),
    ("EBIT ($M)",                "EBIT_M",                '#,##0.0',  True),
    ("Net Income ($M)",          "Net_Income_M",          '#,##0.0',  True),
    ("Net Profit Margin (%)",    "Net_Profit_Margin_Pct", '0.0"%"',   False),
    ("Cost-to-Income Ratio (%)", "Cost_Income_Ratio_Pct", '0.0"%"',   False),
    ("Return on Equity (%)",     "ROE_Pct",               '0.0"%"',   False),
    ("Return on Assets (%)",     "ROA_Pct",               '0.00"%"',  False),
    ("Earnings Per Share ($)",   "EPS",                   '$#,##0.00',False),
]

for ri, (label, col, fmt, bold) in enumerate(metrics, 6):
    label_cell(ws1, ri, 1, label, bold=bold)
    yr_vals = []
    for ci, yr in enumerate(YEARS, 2):
        val = df_a.loc[df_a.Year==yr, col].values[0]
        yr_vals.append(val)
        data_cell(ws1, ri, ci, val, fmt=fmt, bold=bold)
    # YoY change
    yoy = yr_vals[-1] - yr_vals[-2]
    yoy_pct = (yoy / yr_vals[-2]) * 100
    clr = GREEN if yoy >= 0 else RED
    data_cell(ws1, ri, 7, round(yoy_pct, 1), fmt='0.0"%"', color=clr, bold=True)
    # 5Y CAGR
    if yr_vals[0] > 0:
        cagr = ((yr_vals[-1]/yr_vals[0])**(1/4) - 1)*100
        data_cell(ws1, ri, 8, round(cagr, 1), fmt='0.0"%"', bold=True)

set_col_widths(ws1, {"A":30,"B":11,"C":11,"D":11,"E":11,"F":11,"G":10,"H":10})
set_row_height(ws1, range(5, 20), 18)

# ── Revenue Bar Chart ──────────────────────────────────────────────────────────
section_title(ws1, 20, 1, "📊  Revenue & Net Income Trend (2020–2024)", span=8)
for ri, yr in enumerate(YEARS, 22):
    ws1.cell(row=ri, column=1, value=str(yr))
    ws1.cell(row=ri, column=2, value=df_a.loc[df_a.Year==yr,"Total_Revenue_M"].values[0])
    ws1.cell(row=ri, column=3, value=df_a.loc[df_a.Year==yr,"Net_Income_M"].values[0])

chart1 = BarChart()
chart1.type = "col"
chart1.grouping = "clustered"
chart1.title = "Revenue vs Net Income (USD M)"
chart1.y_axis.title = "USD Millions"
chart1.x_axis.title = "Year"
chart1.style = 10
chart1.width = 18
chart1.height = 11

data_ref   = Reference(ws1, min_col=2, max_col=3, min_row=21, max_row=26)
cats_ref   = Reference(ws1, min_col=1, min_row=22, max_row=26)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.series[0].graphicalProperties.solidFill = ACCENT
chart1.series[1].graphicalProperties.solidFill = GREEN
ws1.add_chart(chart1, "A28")

# ── Margin Line Chart ──────────────────────────────────────────────────────────
section_title(ws1, 20, 9, "📈  Profitability Ratios (%)", span=5)
for ri, yr in enumerate(YEARS, 22):
    ws1.cell(row=ri, column=9, value=str(yr))
    ws1.cell(row=ri, column=10, value=df_a.loc[df_a.Year==yr,"Net_Profit_Margin_Pct"].values[0])
    ws1.cell(row=ri, column=11, value=df_a.loc[df_a.Year==yr,"ROE_Pct"].values[0])
    ws1.cell(row=ri, column=12, value=df_a.loc[df_a.Year==yr,"ROA_Pct"].values[0])

chart2 = LineChart()
chart2.title = "Profitability KPIs (%)"
chart2.y_axis.title = "Percentage (%)"
chart2.x_axis.title = "Year"
chart2.style = 10
chart2.width = 14
chart2.height = 11

d2 = Reference(ws1, min_col=10, max_col=12, min_row=21, max_row=26)
c2 = Reference(ws1, min_col=9, min_row=22, max_row=26)
chart2.add_data(d2, titles_from_data=True)
chart2.set_categories(c2)
ws1.add_chart(chart2, "I28")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Monthly Revenue 2024
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Monthly Revenue 2024")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:G2")
t = ws2["A1"]
t.value = "Monthly Revenue & Profitability — FY 2024"
t.font = Font(bold=True, size=15, name="Arial", color=WHITE)
t.fill = PatternFill("solid", fgColor=MID_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

hdrs2 = ["Month","Revenue ($M)","Expenses ($M)","Provision ($M)","Net Income ($M)","Margin (%)","MoM Δ (%)"]
for ci, h in enumerate(hdrs2, 1):
    hdr(ws2, 4, ci, h)

prev_rev = None
for ri, row in enumerate(df_m.itertuples(), 5):
    label_cell(ws2, ri, 1, row.Month)
    data_cell(ws2, ri, 2, row.Revenue_M,    '#,##0.0')
    data_cell(ws2, ri, 3, row.Expenses_M,   '#,##0.0')
    data_cell(ws2, ri, 4, row.Provision_M,  '#,##0.0')
    data_cell(ws2, ri, 5, row.Net_Income_M, '#,##0.0')
    data_cell(ws2, ri, 6, row.Net_Margin_Pct, '0.0"%"')
    if prev_rev:
        mom = ((row.Revenue_M - prev_rev)/prev_rev)*100
        clr = GREEN if mom >= 0 else RED
        data_cell(ws2, ri, 7, round(mom,1), '0.0"%"', color=clr, bold=True)
    else:
        ws2.cell(row=ri, column=7, value="—")
    prev_rev = row.Revenue_M

# Totals row
tr = len(df_m) + 5
label_cell(ws2, tr, 1, "TOTAL / AVG", bold=True)
for ci, col in enumerate(["Revenue_M","Expenses_M","Provision_M","Net_Income_M"], 2):
    data_cell(ws2, tr, ci, f"=SUM({get_column_letter(ci)}5:{get_column_letter(ci)}{tr-1})",
              '#,##0.0', bold=True)
data_cell(ws2, tr, 6, f"=AVERAGE(F5:F{tr-1})", '0.0"%"', bold=True)

# Monthly chart
for ri, row in enumerate(df_m.itertuples(), 20):
    ws2.cell(row=ri, column=1, value=row.Month)
    ws2.cell(row=ri, column=2, value=row.Revenue_M)
    ws2.cell(row=ri, column=3, value=row.Net_Income_M)

chart3 = BarChart()
chart3.type = "col"
chart3.title = "Monthly Revenue vs Net Income — 2024"
chart3.y_axis.title = "USD Millions"
chart3.style = 10; chart3.width = 22; chart3.height = 12
d3 = Reference(ws2, min_col=2, max_col=3, min_row=19, max_row=31)
c3 = Reference(ws2, min_col=1, min_row=20, max_row=31)
chart3.add_data(d3, titles_from_data=True)
chart3.set_categories(c3)
chart3.series[0].graphicalProperties.solidFill = ACCENT
chart3.series[1].graphicalProperties.solidFill = GOLD
ws2.add_chart(chart3, "A35")

set_col_widths(ws2, {"A":12,"B":14,"C":14,"D":14,"E":14,"F":12,"G":12})

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – Asset Quality
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Asset Quality")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:H2")
t3 = ws3["A1"]
t3.value = "Asset Quality & Balance Sheet Metrics — 2020 to 2024"
t3.font = Font(bold=True, size=15, name="Arial", color=WHITE)
t3.fill = PatternFill("solid", fgColor=DARK_BLUE)
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

ah = ["Year","Total Assets ($M)","Total Loans ($M)","Deposits ($M)",
      "NPL Ratio (%)","Capital Adequacy (%)","Liquidity Ratio (%)","Loan/Deposit (%)"]
for ci, h in enumerate(ah, 1):
    hdr(ws3, 4, ci, h)

for ri, row in enumerate(df_as.itertuples(), 5):
    label_cell(ws3, ri, 1, str(int(row.Year)))
    data_cell(ws3, ri, 2, row.Total_Assets_M,     '#,##0.0')
    data_cell(ws3, ri, 3, row.Total_Loans_M,      '#,##0.0')
    data_cell(ws3, ri, 4, row.Total_Deposits_M,   '#,##0.0')
    npl_clr = RED if row.NPL_Ratio_Pct > 2.5 else (GOLD if row.NPL_Ratio_Pct > 1.8 else GREEN)
    data_cell(ws3, ri, 5, row.NPL_Ratio_Pct,      '0.00"%"', color=npl_clr, bold=True)
    cap_clr = GREEN if row.Capital_Adequacy_Pct > 14 else GOLD
    data_cell(ws3, ri, 6, row.Capital_Adequacy_Pct,'0.0"%"', color=cap_clr, bold=True)
    data_cell(ws3, ri, 7, row.Liquidity_Ratio_Pct, '0.0"%"')
    data_cell(ws3, ri, 8, row.Loan_to_Deposit_Pct, '0.0"%"')

# Asset trend chart
for ri, row in enumerate(df_as.itertuples(), 13):
    ws3.cell(row=ri, column=1, value=str(int(row.Year)))
    ws3.cell(row=ri, column=2, value=row.Total_Assets_M)
    ws3.cell(row=ri, column=3, value=row.Total_Loans_M)
    ws3.cell(row=ri, column=4, value=row.Total_Deposits_M)

chart4 = LineChart()
chart4.title = "Asset Growth Trend (USD Millions)"
chart4.y_axis.title = "USD Millions"
chart4.style = 10; chart4.width = 20; chart4.height = 11
d4 = Reference(ws3, min_col=2, max_col=4, min_row=12, max_row=17)
c4 = Reference(ws3, min_col=1, min_row=13, max_row=17)
chart4.add_data(d4, titles_from_data=True)
chart4.set_categories(c4)
ws3.add_chart(chart4, "A20")

set_col_widths(ws3, {"A":8,"B":16,"C":16,"D":14,"E":14,"F":16,"G":16,"H":14})

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – Segment Revenue
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Business Segments")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:F2")
t4 = ws4["A1"]
t4.value = "Business Segment Revenue Analysis — 2020 to 2024"
t4.font = Font(bold=True, size=15, name="Arial", color=WHITE)
t4.fill = PatternFill("solid", fgColor=MID_BLUE)
t4.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 28

# Pivot segment data
pivot = df_s.pivot_table(index="Segment", columns="Year", values="Revenue_M", aggfunc="sum")
pivot = pivot.reset_index()

sh = ["Business Segment"] + [str(y) for y in YEARS] + ["5Y Total","Share (%)"]
for ci, h in enumerate(sh, 1):
    hdr(ws4, 4, ci, h)

total_2024 = df_s[df_s.Year==2024]["Revenue_M"].sum()
for ri, row in enumerate(pivot.itertuples(), 5):
    label_cell(ws4, ri, 1, row.Segment, bold=True)
    row_sum = 0
    for ci, yr in enumerate(YEARS, 2):
        val = getattr(row, f"_{yr}" if str(yr)[0].isdigit() else str(yr), 0)
        # get from pivot correctly
        val = pivot.loc[pivot.Segment==row.Segment, yr].values[0]
        data_cell(ws4, ri, ci, round(val,1), '#,##0.0')
        row_sum += val
    data_cell(ws4, ri, 7, round(row_sum,1), '#,##0.0', bold=True)
    share_2024 = pivot.loc[pivot.Segment==row.Segment, 2024].values[0] / total_2024 * 100
    data_cell(ws4, ri, 8, round(share_2024,1), '0.0"%"', bold=True)

# Pie chart for 2024 segment share
segs_2024 = df_s[df_s.Year==2024].sort_values("Revenue_M", ascending=False)
for ri, row in enumerate(segs_2024.itertuples(), 15):
    ws4.cell(row=ri, column=1, value=row.Segment)
    ws4.cell(row=ri, column=2, value=row.Revenue_M)

pie = PieChart()
pie.title = "2024 Revenue by Business Segment"
pie.style = 10; pie.width = 16; pie.height = 12
pd_ref = Reference(ws4, min_col=2, min_row=14, max_row=19)
pl_ref = Reference(ws4, min_col=1, min_row=15, max_row=19)
pie.add_data(pd_ref, titles_from_data=True)
pie.set_categories(pl_ref)
pie.dataLabels = None
ws4.add_chart(pie, "A22")

set_col_widths(ws4, {"A":22,"B":11,"C":11,"D":11,"E":11,"F":11,"G":11,"H":11})

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 – Quarterly KPIs
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Quarterly KPIs")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:G2")
t5 = ws5["A1"]
t5.value = "Quarterly KPI Scorecard — FY 2024"
t5.font = Font(bold=True, size=15, name="Arial", color=WHITE)
t5.fill = PatternFill("solid", fgColor=DARK_BLUE)
t5.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 28

qh = ["Quarter","Revenue ($M)","Expenses ($M)","Net Income ($M)","EPS ($)","ROE (%)","Exp. Ratio (%)"]
for ci, h in enumerate(qh, 1):
    hdr(ws5, 4, ci, h)

for ri, row in enumerate(df_q.itertuples(), 5):
    label_cell(ws5, ri, 1, row.Quarter, bold=True)
    data_cell(ws5, ri, 2, row.Revenue_M,    '#,##0.0')
    data_cell(ws5, ri, 3, row.Expenses_M,   '#,##0.0')
    data_cell(ws5, ri, 4, row.Net_Income_M, '#,##0.0')
    data_cell(ws5, ri, 5, row.EPS,          '$#,##0.00')
    data_cell(ws5, ri, 6, row.ROE_Pct,      '0.0"%"')
    exp_ratio = (row.Expenses_M / row.Revenue_M) * 100
    clr = GREEN if exp_ratio < 57 else (GOLD if exp_ratio < 60 else RED)
    data_cell(ws5, ri, 7, round(exp_ratio,1), '0.0"%"', color=clr, bold=True)

# Full-year totals
tr5 = 9
label_cell(ws5, tr5, 1, "FULL YEAR 2024", bold=True)
for ci, col in enumerate(["B","C","D"], 2):
    ws5.cell(row=tr5, column=ci, value=f"=SUM({col}5:{col}8)")
    ws5.cell(row=tr5, column=ci).number_format = '#,##0.0'
    ws5.cell(row=tr5, column=ci).font = Font(bold=True, name="Arial")
ws5.cell(row=tr5, column=5, value=f"=SUM(E5:E8)")
ws5.cell(row=tr5, column=5).number_format = '$#,##0.00'
ws5.cell(row=tr5, column=5).font = Font(bold=True, name="Arial")
ws5.cell(row=tr5, column=6, value=f"=AVERAGE(F5:F8)")
ws5.cell(row=tr5, column=6).number_format = '0.0"%"'
ws5.cell(row=tr5, column=6).font = Font(bold=True, name="Arial")

# Quarterly bar chart
for ri, row in enumerate(df_q.itertuples(), 14):
    ws5.cell(row=ri, column=1, value=row.Quarter)
    ws5.cell(row=ri, column=2, value=row.Revenue_M)
    ws5.cell(row=ri, column=3, value=row.Expenses_M)
    ws5.cell(row=ri, column=4, value=row.Net_Income_M)

chart5 = BarChart()
chart5.type = "col"
chart5.grouping = "clustered"
chart5.title = "Quarterly P&L Overview — 2024"
chart5.y_axis.title = "USD Millions"
chart5.style = 10; chart5.width = 20; chart5.height = 11
d5 = Reference(ws5, min_col=2, max_col=4, min_row=13, max_row=17)
c5 = Reference(ws5, min_col=1, min_row=14, max_row=17)
chart5.add_data(d5, titles_from_data=True)
chart5.set_categories(c5)
chart5.series[0].graphicalProperties.solidFill = ACCENT
chart5.series[1].graphicalProperties.solidFill = RED
chart5.series[2].graphicalProperties.solidFill = GREEN
ws5.add_chart(chart5, "A20")

set_col_widths(ws5, {"A":14,"B":14,"C":14,"D":14,"E":12,"F":12,"G":14})

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 – Assumptions & Notes
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Assumptions & Notes")
ws6.sheet_view.showGridLines = False

ws6.merge_cells("A1:D2")
t6 = ws6["A1"]
t6.value = "Model Assumptions, Color Coding & Data Sources"
t6.font = Font(bold=True, size=14, name="Arial", color=WHITE)
t6.fill = PatternFill("solid", fgColor=DARK_BLUE)
t6.alignment = Alignment(horizontal="center", vertical="center")

notes = [
    (5, "COLOR CODING LEGEND", None, True),
    (6, "Blue text", "Hardcoded input assumptions (user-editable)", False),
    (7, "Black text", "Excel formulas and calculated values", False),
    (8, "Green text", "Links pulling from other worksheets", False),
    (9, "Red text", "External links to other files", False),
    (10, "Yellow background", "Key cells requiring attention", False),
    (12, "MODEL ASSUMPTIONS", None, True),
    (13, "Tax Rate", "21.0% (US corporate tax rate)", False),
    (14, "Base Revenue", "$5,200M (FY2020 starting point)", False),
    (15, "Revenue Growth", "Modeled as 6–13% CAGR per year", False),
    (16, "Cost-to-Income", "Modeled at 54–60% of revenue", False),
    (17, "Provision Rate", "6–10% of revenue", False),
    (18, "Loan/Asset Ratio", "60–66% of total assets", False),
    (19, "Deposit/Asset Ratio","70–76% of total assets", False),
    (21, "DATA SOURCES", None, True),
    (22, "All data is synthetic", "Generated for demonstration purposes only", False),
    (23, "Benchmark reference", "Source: FDIC Statistics on Depository Institutions 2024", False),
    (24, "Regulatory thresholds","Capital Adequacy > 10.5% (Basel III CET1)", False),
    (25, "NPL benchmark", "Industry average NPL ratio ~1.5%–2.5% (FDIC 2024)", False),
]

for row_num, label, note, is_title in notes:
    if is_title:
        ws6.cell(row=row_num, column=1, value=label).font = Font(bold=True, size=12, name="Arial", color=MID_BLUE)
        ws6.merge_cells(f"A{row_num}:D{row_num}")
    else:
        c1 = ws6.cell(row=row_num, column=1, value=label)
        c1.font = Font(bold=True, name="Arial", size=10)
        if note:
            c2 = ws6.cell(row=row_num, column=2, value=note)
            c2.font = Font(name="Arial", size=10)
            ws6.merge_cells(f"B{row_num}:D{row_num}")

set_col_widths(ws6, {"A":28,"B":55,"C":10,"D":10})

# ── Save ───────────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f"✅ Excel workbook saved: {OUT_PATH}")
print("   Sheets: Executive Summary | Monthly Revenue | Asset Quality | Business Segments | Quarterly KPIs | Assumptions")
